from openpyxl import load_workbook
from openpyxl.chart import PieChart3D,Reference
import re
from bs4 import *
import urllib.request
class SEO:
    def __init__(self,wbk,wsht):
        self.wbk = wbk
        self.wsht = wsht
    def read_sheet_data(self):
        self.data=[]
        col_no =input("Enter column name : ")
        try:
            temp = wsht[col_no]
            for x in temp:
                if(x!=None):
                    self.data.append(x.value.lower())
        except Exception:
            print("please specify correct column name ")
        if(len(self.data)>0):
            return True
        else:
            return False
    def read_url_from_data(self):
        pattern = re.compile('(https://)|(http://)')
        for x in self.data:
            if(pattern.match(x)):
                self.url = x
                self.data.remove(x)
                return True
        return False
    def web_scrapping(self):
        req = urllib.request.Request(self.url, data=None, 
             headers={ 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36' } 
                                      )
        try:
            req = urllib.request.urlopen(req)
        except Exception:
            return False
        else:
            soup = BeautifulSoup(req,'lxml')
            for x in soup(['script','style']):
                x.extract()
            soup = soup.getText()
            urlConts=[]
            self.urlContents=[]
            for x in soup.splitlines(): #split('\n')
                if(len(x)>0):
                    x=x.lower()
                    urlConts.append(x.strip())
            for x in urlConts:
                self.urlContents.extend(x.split())
            return True
    def CountFrequency(self):
        self.count =[y for y in map((lambda y:self.urlContents.count(y)),self.data)]
        dict ={x:y for x,y in zip(self.data,self.count)}
        print(dict)
    def WriteFrequency(self,book):
        self.wsht['C1'] = 'Words'
        self.wsht['D1'] = 'Frequency'
        i=2
        for x in self.data:
                self.wsht.cell(row = i,column = 3).value = x
                i=i+1
        i=2
        for y in self.count:
                self.wsht.cell(row = i,column = 4).value = y
                i=i+1
        try:
            self.wbk.save(book)
        except Exception:
            print("please check whether the specified file is closed and try again")
            quit()
    def DrawPieChart3D(self,book):
        pie = PieChart3D()
        labels = Reference(self.wsht, min_col=3, min_row=2, max_row=len(self.data))
        data = Reference(self.wsht, min_col=4, min_row=1, max_row=len(self.data))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Pies sold by category"
        self.wsht.add_chart(pie,'G1')
        self.wbk.save(book)
wbk = input("Enter location of workbook : ")
book=wbk
try:
    wbk = load_workbook(wbk)
except Exception:
    print("Sorry unable to locate the file specified")
    print("Hint: use extension (eg .xlsx)")
else:
    wsht = input("Enter worksheet name : ")
    try:
        wsht = wbk[wsht]
    except Exception:
        print("Please enter valid worksheet name")
    else:
        seo = SEO(wbk,wsht)
        if(seo.read_sheet_data()):
            if(seo.read_url_from_data()):
                if(seo.web_scrapping()):
                    seo.CountFrequency()
                    seo.WriteFrequency(book)
                    seo.DrawPieChart3D(book)
                    print("\n"*5)
                    print("Thank you for using our software!")
                    print("Hope you liked it!")
                else:
                    print("Sorry we are unable to connect!")
            else:
                print("No url found")
        else:
            print("The specified column is empty!")